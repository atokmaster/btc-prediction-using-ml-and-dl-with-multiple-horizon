import os
import json
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tensorflow.keras.models import load_model
import joblib
from sklearn.preprocessing import StandardScaler

# === Path Lokal ===
base_dir = os.getcwd()
loop_dir = os.path.join(base_dir, 'loop_results')
realtime_dir = os.path.join(base_dir, 'btc_realtime')
output_path = os.path.join(base_dir, 'signal_log.xlsx')
history_csv = os.path.join(base_dir, 'signal_log_history.csv')
history_xlsx = os.path.join(base_dir, 'signal_log_history.xlsx')

# === Horizon & Mapping
horizon_map = {
    '6H': 'final_comparison_6H.csv',
    '12H': 'final_comparison_12H.csv',
    '1D': 'final_comparison_1D.csv',
    '3D': 'final_comparison_3D.csv',
    '7D': 'final_comparison_7D.csv',
}
source_to_interval = {
    '1H': '1h',
    '4H': '4h',
    '1D': '1d'
}

# === Styling warna di Excel
def apply_excel_styling(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    color_map = {
        'LONG': '92D050',
        'SHORT': 'FF0000',
        'NO TRADE': 'C9C9C9'
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value)
            if val in color_map:
                cell.fill = PatternFill(start_color=color_map[val], end_color=color_map[val], fill_type='solid')
    wb.save(filepath)

# === Proses prediksi
row = {}

for hz, file_name in horizon_map.items():
    try:
        final_df_path = os.path.join(loop_dir, f"Horizon_{hz}", file_name)
        if not os.path.exists(final_df_path):
            raise FileNotFoundError(f"{final_df_path} tidak ditemukan")

        df = pd.read_csv(final_df_path)
        best = df.sort_values('F1_Macro', ascending=False).iloc[0]

        # 🔧 FIX: gunakan langsung Model_Path karena sudah lengkap
        model_path = os.path.join(loop_dir, best['Model_Path'])
        model_type = best['Model_Type']
        source = best['Source']
        interval = source_to_interval[source]

        realtime_path = os.path.join(realtime_dir, f"btc_{interval}_latest.xlsx")
        df_feat = pd.read_excel(realtime_path)

        # Load konfigurasi fitur
        model_folder = os.path.dirname(model_path)
        feat_json_path = os.path.join(model_folder, 'features.json')
        if not os.path.exists(feat_json_path):
            raise FileNotFoundError(f"features.json tidak ditemukan di {model_folder}")
        with open(feat_json_path) as f:
            feat_config = json.load(f)
        feature_cols = feat_config['features']
        feature_len = feat_config['feature_len']

        # Validasi data cukup
        if len(df_feat) < feature_len:
            raise ValueError(f"[{hz}] Data hanya {len(df_feat)} baris, butuh {feature_len}")

        df_input = df_feat[feature_cols]
        scaled = StandardScaler().fit_transform(df_input)
        X_raw = scaled[-feature_len:]

        # Prediksi
        if model_type == 'DL' or model_path.endswith('.keras'):
            model = load_model(model_path)
            X_input = np.expand_dims(X_raw, axis=0)
            probs = model.predict(X_input)[0]
        elif model_type == 'ML' or model_path.endswith('.pkl'):
            model = joblib.load(model_path)
            X_input = StandardScaler().fit_transform(df_input[feature_cols])[-1:]
            probs = model.predict_proba(X_input)[0]
        else:
            raise ValueError(f"Model type tidak dikenali: {model_type}")

        label_map = {0: 'SHORT', 1: 'NO TRADE', 2: 'LONG'}
        pred_label = int(np.argmax(probs))
        confidence = round(float(np.max(probs)), 4)
        gap = round(np.sort(probs)[-1] - np.sort(probs)[-2], 4)

        row[f'Signal_{hz}'] = label_map[pred_label]
        row[f'Conf_{hz}'] = confidence
        row[f'ConfGap_{hz}'] = gap

        print(f"✅ {hz} → {label_map[pred_label]} (conf: {confidence}, gap: {gap})")

    except Exception as e:
        print(f"⚠️ Gagal memproses {hz}: {e}")
        row[f'Signal_{hz}'] = 'ERROR'
        row[f'Conf_{hz}'] = 0.0
        row[f'ConfGap_{hz}'] = 0.0

# === Ensemble Decision
row['Datetime'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
votes = [row[k] for k in row if k.startswith('Signal_') and row[k] != 'ERROR']
if votes:
    vc = pd.Series(votes).value_counts()
    row['Ensemble_Decision'] = vc.idxmax()
    row['Ensemble_Confidence'] = round(vc.max() / len(votes), 4)
else:
    row['Ensemble_Decision'] = 'ERROR'
    row['Ensemble_Confidence'] = 0.0

# === Simpan hasil ke Excel
cols_order = ['Datetime'] + [k for k in row if k != 'Datetime']
df_output = pd.DataFrame([row])[cols_order]
df_output.to_excel(output_path, index=False)
apply_excel_styling(output_path)

# === Simpan ke History CSV & XLSX
if os.path.exists(history_csv):
    df_hist = pd.read_csv(history_csv)
    missing_cols = [col for col in df_output.columns if col not in df_hist.columns]
    for col in missing_cols:
        df_hist[col] = None
    df_hist = pd.concat([df_hist, df_output[df_hist.columns]], ignore_index=True, sort=False)
else:
    df_hist = df_output

df_hist.to_csv(history_csv, index=False)
df_hist.to_excel(history_xlsx, index=False)
apply_excel_styling(history_xlsx)


print("✅ Signal log, CSV, dan Excel history berhasil diperbarui!")
